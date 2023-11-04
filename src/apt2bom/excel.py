"""
Write package lists as Excel sheet.
"""
import os
import logging
from openpyxl import Workbook
from .apt_data import Package, Source
from .resolve_lists import PackageLists
from .output import create_out_dir


logger = logging.getLogger('excel')

source_headers = [
    'package',
    'pkg_type',
    'format',
    'binaries',
    'architecture',
    'version',
    'priority',
    'section',
    'maintainer',
    'standards_version',
    'build_depends',
    'homepage',
    'vcs_browser',
    'vcs_git',
    'directory',
    'package_list',
    'files',
    'url',
    'origin',
    'label',
    'suite',
    'repository_version',
    'codename',
    'date',
    'description',
    'component'
]

package_headers = [
    'package',
    'pkg_type',
    'architecture',
    'version',
    'source_package',
    'priority',
    'section',
    'maintainer',
    'original_maintainer',
    'bugs',
    'installed_size',
    'provides',
    'depends',
    'recommends',
    'suggests',
    'filename',
    'size',
    'md5',
    'sha1',
    'sha256',
    'sha512',
    'homepage',
    'description',
    'task',
    'description_md5',
    'url',
    'origin',
    'label',
    'suite',
    'repository_version',
    'codename',
    'date',
    'repository_description',
    'component'
    'source_format',
    'source_binaries',
    'source_architecture',
    'source_version',
    'source_priority',
    'source_section',
    'source_maintainer',
    'source_standards_version',
    'source_build_depends',
    'source_homepage',
    'source_vcs_browser',
    'source_vcs_git',
    'source_directory',
    'source_package_list',
    'source_files',
    'source_url',
    'source_origin',
    'source_label',
    'source_suite',
    'source_repository_version',
    'source_codename',
    'source_date',
    'source_description',
    'source_component'
]


def write_excel_package_list(config, lists: PackageLists):
    """
    Write package lists as Excel file.
    """

    create_out_dir(config)


    wb = Workbook()
    
    # write sources list
    ws = wb.active
    ws.title = 'Source ECU Packages'
    sources: dict[str, Source] = {}
    # collect all source packages
    for arch in config['packages']['architectures']:
        for name in lists.ecu_packages.keys():
            if arch in lists.ecu_packages[name]:
                package = lists.ecu_packages[name][arch]
                if package.source:
                    if package.source.package not in sources:
                        package.source.pkg_type = package.pkg_type
                        sources[package.source.package] = package.source

    source_list = sources.values()
    source_list.sort(key=lambda source: source.package)

    ws.append(source_headers)
        
    for i, source in enumerate(source_list):
        data = source.to_record()
        for j, key in enumerate(source_headers):
            if key in data:
                ws.cell(row=i+2, column=j+1, value=data[key])
    
    # write architecture specific package lists
    for arch in config['packages']['architectures']:
        ws = wb.create_sheet(f'{arch} ECU Packages')
        ecu_packages : list[Package] = []
        for name in lists.ecu_packages:
            if arch in lists.ecu_packages[name]:
                ecu_packages.append(lists.ecu_packages[name][arch])

        if ecu_packages:
            ecu_packages.sort(key=lambda package: package.package)

            ws.append(package_headers)
        
            for i, package in enumerate(ecu_packages):
                data = package.to_record()
                for j, key in enumerate(package_headers):
                    if key in data:
                        ws.cell(row=i+2, column=j+1, value=data[key])
        
        ws = wb.create_sheet(f'{arch} SDK Packages')

        sdk_packages : list[Package] = []
        for name in lists.sdk_packages:
            if arch in lists.sdk_packages[name]:
                sdk_packages.append(lists.sdk_packages[name][arch])

        if sdk_packages:
            sdk_packages.sort(key=lambda package: package.package)

            ws.append(package_headers)
            
            for i, package in enumerate(sdk_packages):
                data = package.to_record()
                for j, key in enumerate(package_headers):
                    if key in data:
                        ws.cell(row=i+2, column=j+1, value=data[key])

    file = os.path.join(config['output']['directory'], config['output']['excel'])
    wb.save(file)
